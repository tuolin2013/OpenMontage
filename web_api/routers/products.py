"""
products.py — GET /api/products
从 data/products.xlsx 读取商品列表，返回给前端供下拉选择自动填充。
"""
from pathlib import Path
from fastapi import APIRouter, HTTPException

router = APIRouter(prefix="/products", tags=["products"])

XLSX_PATH = Path(__file__).resolve().parents[2] / "data" / "products.xlsx"

COLUMNS = [
    "产品名称", "品牌", "品类", "规格", "价格",
    "核心成分", "产品卖点", "使用说明", "注意事项", "产品实拍图",
]


def _load_products():
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    if not XLSX_PATH.exists():
        raise HTTPException(404, f"products.xlsx not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Header row
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Map column name → index
    idx = {h: i for i, h in enumerate(headers)}

    def cell(row, col_name):
        i = idx.get(col_name)
        if i is None:
            return None
        cells = list(row)
        return cells[i].value if i < len(cells) else None

    products = []
    for row in ws.iter_rows(min_row=2):
        name = cell(row, "产品名称")
        if not name:
            continue
        products.append({
            "name":            str(name).strip(),
            "brand":           str(cell(row, "品牌") or "").strip(),
            "category":        str(cell(row, "品类") or "").strip(),
            "spec":            str(cell(row, "规格") or "").strip(),
            "price":           cell(row, "价格"),
            "key_ingredients": str(cell(row, "核心成分") or "").strip(),
            "selling_points":  str(cell(row, "产品卖点") or "").strip(),
            "usage":           str(cell(row, "使用说明") or "").strip(),
            "notes":           str(cell(row, "注意事项") or "").strip(),
            "image_url":       str(cell(row, "产品实拍图") or "").strip(),
        })

    wb.close()
    return products


@router.get("")
def list_products():
    """返回所有商品列表（轻量版，含 name / brand / image_url 用于下拉）"""
    return {"products": _load_products()}


@router.get("/{name}")
def get_product(name: str):
    """根据产品名称精确查找，返回完整字段供表单自动填充"""
    products = _load_products()
    for p in products:
        if p["name"] == name:
            return p
    raise HTTPException(404, f"Product '{name}' not found")
